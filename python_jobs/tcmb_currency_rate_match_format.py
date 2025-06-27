
import requests
from datetime import datetime, timedelta
from xml.etree import ElementTree as ET
from openpyxl import Workbook, load_workbook
import os

def fetch_tcmb_rate(date):
    try:
        url = f"https://www.tcmb.gov.tr/kurlar/{date.strftime('%Y%m')}/{date.strftime('%d%m%Y')}.xml"
        response = requests.get(url)
        if response.status_code != 200:
            return None
        tree = ET.fromstring(response.content)
        for currency in tree.findall("Currency"):
            if currency.attrib.get("Kod") == "EUR":
                inverse = currency.find("ForexBuying").text
                inverse = float(inverse)
                company_rate = round(1 / inverse, 6)
                return {
                    "Date": date.strftime("%Y-%m-%d"),
                    "Company Rate": company_rate,
                    "Inverse Company Rate": round(inverse, 5)
                }
    except Exception:
        return None

def get_user_dates():
    print("📆 Please enter the date range to fetch historical EUR/TRY rates from TCMB.")
    start_input = input("Start date (YYYY-MM-DD): ")
    end_input = input("End date (YYYY-MM-DD, leave blank for today): ")

    try:
        start_date = datetime.strptime(start_input, "%Y-%m-%d")
        end_date = datetime.strptime(end_input, "%Y-%m-%d") if end_input else datetime.now()
    except ValueError:
        print("❌ Invalid date format. Please use YYYY-MM-DD.")
        exit(1)

    if start_date > end_date:
        print("❌ Start date must be before end date.")
        exit(1)

    return start_date, end_date

def save_monthly_excel(data, year, month):
    filename = f"Currency_Rate_{year}_{month:02}.xlsx"
    append = os.path.exists(filename)
    
    if append:
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Currency Rate"
        ws.append(["Date", "Company Rate", "Inverse Company Rate"])

    for row in data:
        ws.append([
            row["Date"],
            row["Company Rate"],
            row["Inverse Company Rate"]
        ])

    wb.save(filename)
    print(f"💾 Saved {len(data)} records to '{filename}'")

def main():
    start_date, end_date = get_user_dates()
    print(f"🔍 Fetching rates from {start_date.date()} to {end_date.date()}...\n")

    current_date = start_date
    batch_data = []
    batch_size = 30

    while current_date <= end_date:
        if current_date.weekday() < 5:
            print(f"📅 Processing {current_date.strftime('%Y-%m-%d')}...", end=' ')
            result = fetch_tcmb_rate(current_date)
            if result:
                batch_data.append(result)
                print(f"✅ {result['Inverse Company Rate']}")
            else:
                print("⚠️ No rate found")
        else:
            print(f"🛌 Skipping weekend: {current_date.strftime('%Y-%m-%d')}")

        next_date = current_date + timedelta(days=1)
        if (len(batch_data) >= batch_size or
            (next_date.month != current_date.month or next_date > end_date)) and batch_data:

            save_monthly_excel(batch_data, current_date.year, current_date.month)
            batch_data.clear()

        current_date += timedelta(days=1)

    print("\n✅ All done! Files saved for each month.")

if __name__ == "__main__":
    main()
