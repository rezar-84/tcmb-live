
import requests
from datetime import datetime, timedelta
from xml.etree import ElementTree as ET
from openpyxl import Workbook, load_workbook
import os

def fetch_tcmb_rate(date, currency_code):
    try:
        url = "https://www.tcmb.gov.tr/kurlar/{}/{}.xml".format(
            date.strftime('%Y%m'), date.strftime('%d%m%Y'))
        response = requests.get(url)
        if response.status_code != 200:
            return None
        tree = ET.fromstring(response.content)
        for currency in tree.findall("Currency"):
            if currency.attrib.get("Kod") == currency_code:
                inverse = currency.find("ForexBuying").text
                inverse = float(inverse)
                company_rate = round(1 / inverse, 6)
                return {
                    "Date": date.strftime("%Y-%m-%d"),
                    "Company Rate": company_rate,
                    "Inverse Company Rate": round(inverse, 5)
                }
    except Exception:
        return None

def get_user_input():
    print("📆 Enter date range to fetch historical currency rates from TCMB.")
    start_input = input("Start date (YYYY-MM-DD): ")
    end_input = input("End date (YYYY-MM-DD, leave blank for today): ")
    currency_code = input("Currency code (e.g., EUR, USD): ").upper()
    save_mode = input("Save mode - monthly or yearly? (default: monthly): ").strip().lower() or "monthly"

    try:
        start_date = datetime.strptime(start_input, "%Y-%m-%d")
        end_date = datetime.strptime(end_input, "%Y-%m-%d") if end_input else datetime.now()
    except ValueError:
        print("❌ Invalid date format. Please use YYYY-MM-DD.")
        exit(1)

    if start_date > end_date:
        print("❌ Start date must be before end date.")
        exit(1)

    return start_date, end_date, currency_code, save_mode

def save_excel(data, folder, year, month=None):
    filename = "Currency_Rate_{}_{}.xlsx".format(year, f"{month:02}" if month else "ALL") if month else "Currency_Rate_{}.xlsx".format(year)
    filepath = os.path.join(folder, filename)
    append = os.path.exists(filepath)

    if append:
        wb = load_workbook(filepath)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Currency Rate"
        ws.append(["Date", "Company Rate", "Inverse Company Rate"])

    for row in data:
        ws.append([
            row["Date"],
            row["Company Rate"],
            row["Inverse Company Rate"]
        ])

    wb.save(filepath)
    print("💾 Saved {} records to '{}'".format(len(data), filepath))

def main():
    start_date, end_date, currency_code, save_mode = get_user_input()
    run_folder = "run_{}".format(datetime.now().strftime('%Y-%m-%d'))
    os.makedirs(run_folder, exist_ok=True)

    print("🔍 Fetching {} rates from {} to {}...\n".format(currency_code, start_date.date(), end_date.date()))

    current_date = start_date
    batch_data = []
    batch_size = 30

    while current_date <= end_date:
        if current_date.weekday() < 5:
            print("📅 Processing {}...".format(current_date.strftime('%Y-%m-%d')), end=' ')
            result = fetch_tcmb_rate(current_date, currency_code)
            if result:
                batch_data.append(result)
                print("✅ {}".format(result['Inverse Company Rate']))
            else:
                print("⚠️ No rate found")
        else:
            print("🛌 Skipping weekend: {}".format(current_date.strftime('%Y-%m-%d')))

        next_date = current_date + timedelta(days=1)
        date_changed = (
            (save_mode == "monthly" and (next_date.month != current_date.month or next_date > end_date)) or
            (save_mode == "yearly" and (next_date.year != current_date.year or next_date > end_date))
        )

        if (len(batch_data) >= batch_size or date_changed) and batch_data:
            if save_mode == "monthly":
                save_excel(batch_data, run_folder, current_date.year, current_date.month)
            else:
                save_excel(batch_data, run_folder, current_date.year)
            batch_data.clear()

        current_date += timedelta(days=1)

    print("\n✅ Done! Files saved in folder:", run_folder)

if __name__ == "__main__":
    main()
