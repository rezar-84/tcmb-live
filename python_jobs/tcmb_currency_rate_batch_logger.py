
import requests
from datetime import datetime, timedelta
from xml.etree import ElementTree as ET
from openpyxl import Workbook, load_workbook
import os

def fetch_tcmb_rate(date):
    try:
        url = f"https://www.tcmb.gov.tr/kurlar/{date.strftime('%Y%m')}/{date.strftime('%d%m%Y')}.xml"
        response = requests.get(url)
        if response.status_code != 200:
            return None
        tree = ET.fromstring(response.content)
        for currency in tree.findall("Currency"):
            if currency.attrib.get("Kod") == "EUR":
                rate = currency.find("ForexBuying").text
                return float(rate)
    except Exception:
        return None

def get_user_dates():
    print("📆 Please enter the date range to fetch historical EUR/TRY rates from TCMB.")
    start_input = input("Start date (YYYY-MM-DD): ")
    end_input = input("End date (YYYY-MM-DD, leave blank for today): ")

    try:
        start_date = datetime.strptime(start_input, "%Y-%m-%d")
        end_date = datetime.strptime(end_input, "%Y-%m-%d") if end_input else datetime.now()
    except ValueError:
        print("❌ Invalid date format. Please use YYYY-MM-DD.")
        exit(1)

    if start_date > end_date:
        print("❌ Start date must be before end date.")
        exit(1)

    return start_date, end_date

def save_to_excel(data, file_path, append=False):
    if append and os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Currency Rate"
        ws.append(["Id", "Currency", "Rate", "Date", "Company"])

    starting_id = ws.max_row if append else 1

    for i, row in enumerate(data):
        ws.append([
            starting_id + i,
            row["Currency"],
            row["Rate"],
            row["Date"],
            row["Company"]
        ])

    wb.save(file_path)

def main():
    start_date, end_date = get_user_dates()
    print(f"🔍 Fetching rates from {start_date.date()} to {end_date.date()}...\n")

    current_date = start_date
    batch_data = []
    batch_size = 30
    file_path = "Currency_Rate_TCMB.xlsx"
    record_counter = 0

    while current_date <= end_date:
        if current_date.weekday() < 5:  # Weekdays only
            print(f"📅 Processing {current_date.strftime('%Y-%m-%d')}...", end=' ')
            rate = fetch_tcmb_rate(current_date)
            if rate:
                record = {
                    "Currency": "EUR",
                    "Rate": rate,
                    "Date": current_date.strftime("%Y-%m-%d 11:00:00"),
                    "Company": "Your Company"
                }
                batch_data.append(record)
                record_counter += 1
                print(f"✅ {rate}")
            else:
                print("⚠️ No rate found")
        else:
            print(f"🛌 Skipping weekend: {current_date.strftime('%Y-%m-%d')}")

        if len(batch_data) >= batch_size:
            save_to_excel(batch_data, file_path, append=True)
            print(f"💾 Saved {len(batch_data)} records to '{file_path}'")
            batch_data.clear()

        current_date += timedelta(days=1)

    if batch_data:
        save_to_excel(batch_data, file_path, append=True)
        print(f"💾 Saved remaining {len(batch_data)} records to '{file_path}'")

    print("\n✅ Done! All rates saved in:", file_path)

if __name__ == "__main__":
    main()
